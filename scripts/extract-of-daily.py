#!/usr/bin/env python3
"""Extract daily OF metrics from the OFM Google Sheets workbook export."""

from __future__ import annotations

import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


MODEL_SHEETS = (
    ("Carla", "@yourfavouritefilipina"),
    ("Jacky", "@pinayqueenb"),
    ("Jaysie", "@jaysie008"),
)
START_DATE = date(2026, 6, 1)
SHEET_ID = "10d9hfGMYy3HupRUG4NgU49ycJogeeBJ6Anqu-jKsuCk"


def as_number(value: object) -> float:
    if value is None:
        return 0.0
    return round(float(value), 2)


def as_jordan_net(value: object) -> float | None:
    if value is None:
        return None
    return round(float(value), 3)


def extract(workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    models: list[dict[str, object]] = []
    available_dates: list[date] = []

    for model_name, handle in MODEL_SHEETS:
        worksheet = workbook[model_name]
        daily: list[dict[str, object]] = []

        for (
            _,
            captured_at,
            earning,
            new_subs,
            fans_count,
            _,
            jordan_net,
        ) in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            max_col=7,
            values_only=True,
        ):
            if not isinstance(captured_at, datetime):
                continue
            captured_date = captured_at.date()
            if captured_date < START_DATE:
                continue
            if (
                earning is None
                and new_subs is None
                and fans_count is None
                and jordan_net is None
            ):
                continue

            available_dates.append(captured_date)
            daily.append(
                {
                    "date": captured_date.isoformat(),
                    "netEarning": as_number(earning),
                    "newSubs": int(new_subs or 0),
                    "fansCount": int(fans_count) if fans_count is not None else None,
                    "jordanNet": as_jordan_net(jordan_net),
                }
            )

        models.append({"name": model_name, "handle": handle, "daily": daily})

    return {
        "source": "OFM - Daily Earnings",
        "sheetId": SHEET_ID,
        "syncedAt": datetime.now(timezone.utc).isoformat(timespec="seconds").replace(
            "+00:00", "Z"
        ),
        "availableFrom": min(available_dates).isoformat(),
        "availableTo": max(available_dates).isoformat(),
        "currency": "USD",
        "models": models,
    }


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract-of-daily.py INPUT.xlsx OUTPUT.json")

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = Path(sys.argv[2]).expanduser().resolve()
    payload = extract(workbook_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
